import os
from core.utils.logger import leer_logs
from core.envios.enviar_correo import enviar_correo
from openpyxl import Workbook
from openpyxl.styles import Font
from dotenv import load_dotenv

load_dotenv()

EMAIL_DESTINO = os.getenv("EMAIL_RESUMEN")



def clasificar_log(nombre_archivo):
    nombre = nombre_archivo.lower()
    if "envio" in nombre:
        return "Envíos"
    elif "factura" in nombre:
        return "Facturación"
    elif "csv" in nombre:
        return "Generación de CSV"
    elif "usuario" in nombre or "user" in nombre:
        return "Usuarios Automáticos"
    else:
        return "Otros Procesos"

def resumen_logs_multiples():
    carpeta_logs = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../data/logs"))
    carpeta_salida = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../data/logs"))

    if not os.path.exists(carpeta_logs):
        print("No se encontró la carpeta de logs.")
        return
    
    os.makedirs(carpeta_salida, exist_ok=True)

    archivos_log = [f for f in os.listdir(carpeta_logs) if f.endswith(".log")]
    if not archivos_log:
        print("No hay archivos .log para procesar.")
        return

    secciones = {}

    for archivo in archivos_log:
        ruta_log = os.path.join(carpeta_logs, archivo)
        registros = leer_logs(ruta_log)
        if not registros:
            continue

        categoria = clasificar_log(archivo)
        if categoria not in secciones:
            secciones[categoria] = []

        for linea in registros:
            if not linea.startswith("#"):
                secciones[categoria].append((archivo, linea.strip()))

    if not secciones:
        print("No se encontró contenido útil en los archivos log.")
        return

  
    wb = Workbook()
    wb.remove(wb.active)  

    for categoria, entradas in secciones.items():
        ws = wb.create_sheet(title=categoria[:31])  

       
        ws.append(["Archivo", "Fecha y Hora", "Mensaje"])
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for archivo, linea in entradas:
            if "|" in linea:
                partes = linea.split("|", 1)
                fecha_hora = partes[0].strip()
                mensaje = partes[1].strip()
            else:
                fecha_hora = ""
                mensaje = linea

            ws.append([archivo, fecha_hora, mensaje])

    
    ruta_excel = os.path.join(carpeta_salida, "resumen_logs_multiples.xlsx")
    wb.save(ruta_excel)

    
    asunto = "📊 Resumen Consolidado de Archivos Log del Sistema"
    cuerpo = (
        "Se adjunta el resumen consolidado en formato Excel de los logs del sistema.\n"
        "Cada hoja corresponde a una categoría (Envíos, Facturación, etc.)."
    )
    resultado = enviar_correo(EMAIL_DESTINO, asunto, cuerpo, archivos_adjuntos=[ruta_excel])
    print(f"Correo enviado: {resultado}")
